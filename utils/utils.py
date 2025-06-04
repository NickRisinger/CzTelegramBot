import pytz
from openpyxl import load_workbook
from datetime import datetime
from config.config import Config

gtins = []

GIFTS = {}


def load_gtins_from_excel():
    workbook = load_workbook(Config.PATH_FILE_ALLOWED_GTINS)
    sheet = workbook.active

    gtins.clear()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] is None:
            gtins.append(row[0])

    print(gtins)


def load_products_from_excel():
    workbook = load_workbook(Config.PATH_FILE_PRODUCTS)
    sheet = workbook.active

    GIFTS.clear()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] is None:
            GIFTS[str(row[0])] = {'name': row[1], 'price': row[2], 'description': row[3], 'alcohol': bool(row[4]), 'image': row[5]}


def format_date(date: str) -> str:
    utc_time_str = date
    utc_dt = datetime.fromisoformat(utc_time_str).replace(tzinfo=pytz.UTC)
    local_dt = utc_dt.astimezone(pytz.timezone("Europe/Moscow"))
    return local_dt.strftime("%H:%M %d-%m-%Y")